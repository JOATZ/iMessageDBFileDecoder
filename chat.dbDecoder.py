# ----------------------------------------
# Script Author: theJOATZ
# Description: This script exports iMessages from an SQLite database (chat.db) to an Excel workbook.
# Each chat is saved in a separate worksheet.
# ----------------------------------------

import sqlite3
import pandas as pd
import openpyxl as px

# Define file paths here
db_file_path = 'C:/Users/Z/Desktop/chat.db'  # Path to the database file
excel_output_path = 'C:/Users/Z/Desktop/iMessages.xlsx'  # Path for the Excel output file

# Connect to the database
conn = sqlite3.connect(db_file_path)

# Get a list of unique handle IDs
handle_ids_query = "SELECT DISTINCT handle_id FROM message"
handle_ids = pd.read_sql_query(handle_ids_query, conn)['handle_id']

# Create a Pandas Excel writer object
with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
    # Iterate through each handle ID and write to a separate sheet
    for handle_id in handle_ids:
        query = f"""
            SELECT
                message.rowid,
                handle.id,
                message.date,
                message.is_from_me,
                message.text
            FROM
                message
            LEFT JOIN
                handle ON message.handle_id = handle.ROWID
            WHERE
                handle_id = {handle_id}
            ORDER BY
                message.date
        """
        data = pd.read_sql_query(query, conn)
        sheet_name = f'Chat_{handle_id}'  # Name the sheet based on the handle ID
        data.to_excel(writer, sheet_name=sheet_name, index=False)

# Close the connection
conn.close()

# Insert a column for Apple Epoch time converison, convert Epoch time
wb = px.load_workbook(excel_output_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    ws.insert_cols(idx=4, amount=1)  # Inserting a new column at index 4 (column D)
    ws['D1'] = 'Formatted Date'  # Adding header for the new column
    
    for row in range(2, ws.max_row + 1):  # Assuming header is in row 1
        date_formula = f'TEXT(DATE(2001,1,1) + (C{row}/1000000000/86400), "yyyy-mm-dd hh:mm:ss")'
        ws.cell(row=row, column=4, value=f'={date_formula}')  # Writing formula to new column

wb.save(excel_output_path)

# Reopen the Excel file to adjust the date column and column widths
wb = px.load_workbook(excel_output_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Adjusting column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 100
    
    # Enabling text wrapping for column F
    for row in ws['F']:
        row.alignment = px.styles.Alignment(wrap_text=True)
    
    # Setting alignment for columns A, B, C, D, and E
    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        for row in ws[col_letter]:
            row.alignment = px.styles.Alignment(horizontal='center', vertical='center')
    # Replace '1' or '1' with 'Me' in column E
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
        for cell in row:
            if str(cell.value) == '1':  # This will match both number 1 and string '1'
                cell.value = 'Me'

wb.save(excel_output_path)
